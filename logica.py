import sqlite3
import json
import pandas as pd
import shutil
import os
import platform
import subprocess
import time
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from fpdf import FPDF

def crear_bd():
    conexion = sqlite3.connect("inventario.db")
    cursor = conexion.cursor()

    cursor.execute("""
    CREATE TABLE IF NOT EXISTS productos (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        codigo TEXT,
        nombre TEXT NOT NULL,
        categoria TEXT NOT NULL,
        precio REAL NOT NULL,
        stock INTEGER NOT NULL
    )
    """)

    conexion.commit()
    conexion.close()

def cargar_datos():
    conexion = sqlite3.connect("inventario.db")
    conexion.row_factory = sqlite3.Row
    cursor = conexion.cursor()

    cursor.execute("SELECT * FROM productos")

    productos = []

    for fila in cursor.fetchall():
        productos.append(dict(fila))

    conexion.close()

    return productos

        #SE PUEDE ESCANEAR EL PRODUCTO USANDO POR EJEMPLO: BARCODE
def agregar_producto(productos):
    print("\n>>> Agregar Producto <<<")
    print("\n" + "-" * 20)
    print("0. Volver Atrás")
    print("-" * 20 + "\n")

    while True:
        print("Escanee el código o presione ENTER si no tiene:")
        codigo = input("Código: ").strip()

        if codigo == "0":
            print("Operación cancelada.")
            return

        if codigo != "":
            existe = any(p.get('codigo') == codigo for p in productos)
            if existe:
                print(f"Error: El código '{codigo}' ya está registrado.")
                continue
        else:
            codigo = "S/C"

        break

    while True:
        nombre = input("Nombre del Producto: ").strip().capitalize()
        if nombre == "0":
            return
        if nombre != "":
            break
        print("Error: El nombre es obligatorio.")

    while True:
        categoria = input("Categoría: ").strip().capitalize()
        if categoria == "0":
            return
        if categoria != "":
            break
        print("Error: La categoría es obligatoria.")

    while True:
        precio_str = input("Precio: ")
        if precio_str == "0":
            return
        try:
            precio_int = int(precio_str)
            if precio_int > 0:
                break
            print("Error: El precio debe ser positivo.")
        except ValueError:
            print("Error: Ingrese un número válido.")

    while True:
        stock_input = input("Stock: ")
        if stock_input.isdigit():
            stock = int(stock_input)
            break
        print("Ingrese una cantidad válida.")

    conexion = sqlite3.connect("inventario.db")
    cursor = conexion.cursor()

    cursor.execute("""
        INSERT INTO productos
        (codigo, nombre, categoria, precio, stock)
        VALUES (?, ?, ?, ?, ?)
    """, (
        codigo,
        nombre,
        categoria,
        precio_int,
        stock
    ))

    conexion.commit()
    conexion.close()

    print(f"\nProducto '{nombre}' agregado correctamente.")
            
def mostrar_inventario(productos):
        if len(productos) == 0:
            print("Inventario vacio")
        else:
            print(f"\n{'N°':<3} | {'Codigo':<15} | {'Nombre':<15} | {'Categoría':<15} | {'Precio':<15} | {'stock':<10}")
            print("-" * 90)
            for i, p in enumerate(productos, start=1):
                    print(f"{i:<3} | {p['codigo']:<15} | {p['nombre']:<15} | {p['categoria']:<15} | ${p['precio']:<14} | {p['stock']}")
            
            
def buscar_producto(productos):
    print("\n" + "="*30)
    print("      BUSCADOR")
    print("="*30)
    
    termino = input("Escanee código o busque por nombre/categoría: ").lower().strip()
    
    if termino == "0" or termino == "":
        return

    encontrados = []

    for p in productos:
        en_codigo = p.get('codigo') == termino
        
        en_nombre = termino in p['nombre'].lower()
        en_categoria = termino in p['categoria'].lower()
        
        if en_codigo or en_nombre or en_categoria:
            encontrados.append(p)

    if encontrados:
        print(f"\nResultados encontrados ({len(encontrados)}):")
        print(f"\n{'N°':<3} | {'Nombre':<15} | {'Categoría':<12} | {'Precio':<10} | {'Stock':<6}")
        print("-" * 60)
        for i, p in enumerate(encontrados, start=1):
            print(f"{i:<3} | {p['nombre']:<15} | {p['categoria']:<12} | ${p['precio']:<9} | {p['stock']}")
    else:
        print(f"\nNo se encontró nada que coincida con '{termino}'")
    
                
def eliminar_producto(productos):
    print("\n>>> Eliminador de productos <<<\n")
    if len(productos) == 0:
        print("No hay productos")
        return

    print("Seleccione el número de producto o escanee el código\n")
    print(f"{'N°':<3} | {'Codigo':<15} | {'Nombre':<15} | {'Categoría':<15} | {'Precio':<10}")
    print("-" * 90)
    for i, p in enumerate(productos, start=1):
        print(f"{i:<3} | {p.get('codigo', 'S/C'):<15} | {p['nombre']:<15} | {p['categoria']:<15} | ${p['precio']:<10}")
        
    print("\n" + "-" * 20)
    print("0. Volver al menú principal")
    print("-" * 20)
        
    entrada = input("\nNúmero de lista o escanee código: ").strip()
    
    if entrada == "0":
        print("Operación cancelada.")
        return

    indice_a_borrar = None
    if entrada.isdigit() and int(entrada) <= len(productos):
        indice_a_borrar = int(entrada) - 1
    else:
        for i, p in enumerate(productos):
            if p.get('codigo') == entrada:
                indice_a_borrar = i
                break

    if indice_a_borrar is not None:
        nombre_producto = productos[indice_a_borrar]['nombre']

        confirmar = input(
            f"\n¿Estás seguro que quieres eliminar {nombre_producto}? (si/no): "
        ).lower()

        if confirmar == 'si':

            producto = productos[indice_a_borrar]

            conexion = sqlite3.connect("inventario.db")
            cursor = conexion.cursor()

            cursor.execute(
                "DELETE FROM productos WHERE id = ?",
                (producto["id"],)
            )

            conexion.commit()
            conexion.close()

            print(f"\nEl producto {producto['nombre']} se eliminó con éxito")

        else:
            print("No se eliminó el producto")

    else:
        print("No se encontró ningún producto con esa enumeración o código")
                    
                    
def generar_reporte(productos):
    print("\n" + "="*30)
    print("   REPORTE DE INVENTARIO")
    print("="*30)
    
    if not productos:
        print("No hay productos para generar estadísticas.")
        return

    total_items = len(productos)
    valor_total_inventario = 0
    total_unidades = 0 
    producto_mas_caro = productos[0]

    for p in productos:
        valor_total_inventario += (p['precio'] * p['stock'])
        total_unidades += p['stock']
        
        if p['precio'] > producto_mas_caro['precio']:
            producto_mas_caro = p

    print(f"Total de productos distintos: {total_items}")
    print(f"Total de unidades en deposito:  {total_unidades}")
    print(f"Valor Total del Inventario:  ${valor_total_inventario}")
    print(f"Producto de mayor precio:    {producto_mas_caro['nombre']} (${producto_mas_caro['precio']})")
    
    print("\n" + "-"*30)
    print("   ALERTAS DE REPOSICIÓN")
    print("!"*30)
    
    productos_bajos = [p for p in productos if p['stock'] < 5]
    
    if productos_bajos:
        print(f"{'PRODUCTO':<20} | {'STOCK ACTUAL':<10}")
        print("-" * 35)
        for p in productos_bajos:
            print(f" {p['nombre']:<18} | Quedan: {p['stock']}")
        print(f"\nTotal: {len(productos_bajos)} productos necesitan stock.")
    else:
        print("Todo en orden. No hay productos con stock bajo.")
        
        
def editar_producto(productos):
    print("\n>>> Modificar Producto <<<\n")

    if len(productos) == 0:
        print("No hay productos para modificar")
        return

    print(f"\n{'N°':<3} | {'Nombre':<15} | {'Categoría':<15} | {'Precio':<10}")
    print("-" * 60)

    for i, p in enumerate(productos, start=1):
        print(
            f"{i:<3} | {p['nombre']:<15} | "
            f"{p['categoria']:<15} | ${p['precio']:<10}"
        )

    print("\n" + "-" * 20)
    print("0. Volver al menú principal")
    print("-" * 20)

    posicion_str = input("\nIngrese el producto a modificar: ").strip()

    if posicion_str == "0":
        print("Operación cancelada.")
        return

    if posicion_str.isdigit():

        indice = int(posicion_str)

        if 1 <= indice <= len(productos):

            producto_actual = productos[indice - 1]

            print(f"\n>>> Editando: {producto_actual['nombre']} <<<\n")

            nuevo_nombre = input(
                f"Nombre Actual ({producto_actual['nombre']}): "
            ).strip()

            if nuevo_nombre == "":
                nuevo_nombre = producto_actual['nombre']

            nueva_categoria = input(
                f"Categoría Actual ({producto_actual['categoria']}): "
            ).strip()

            if nueva_categoria == "":
                nueva_categoria = producto_actual['categoria']

            nuevo_precio_str = input(
                f"Precio Actual (${producto_actual['precio']}): "
            ).strip()

            if nuevo_precio_str == "":
                nuevo_precio = producto_actual['precio']

            elif nuevo_precio_str.isdigit():
                nuevo_precio = int(nuevo_precio_str)

            else:
                print("Precio inválido. Se conservará el valor actual.")
                nuevo_precio = producto_actual['precio']

            conexion = sqlite3.connect("inventario.db")
            cursor = conexion.cursor()

            cursor.execute("""
                UPDATE productos
                SET nombre = ?,
                    categoria = ?,
                    precio = ?
                WHERE id = ?
            """, (
                nuevo_nombre,
                nueva_categoria,
                nuevo_precio,
                producto_actual["id"]
            ))

            conexion.commit()
            conexion.close()

            print("\n✅ Modificación exitosa.\n")

        else:
            print("No existe ese número en el inventario.")

    else:
        print("Debes ingresar el número de fila.")
            
def registrar_venta(productos):
    os.makedirs("backups", exist_ok=True)
    os.makedirs("tickets", exist_ok=True)

    print("\n" + "="*30)
    print("   SISTEMA DE VENTAS")
    print("="*30)
    print("Escanee el código o escriba el nombre.")
    
    carrito = []
    total_venta = 0
    
    while True:
        entrada = input("\nProducto / Escáner (o 'fin' para cobrar): ").strip()
        
        if entrada.lower() == 'fin':
            break

        encontrado = False
        for p in productos:
            if p.get('codigo') == entrada or p['nombre'].lower() == entrada.lower():
                encontrado = True
                print(f"Producto: {p['nombre']} | Precio: ${p['precio']} | Stock: {p['stock']}")
                
                try:
                    cant_input = input("Cantidad a vender: ").strip()
                    cantidad = int(cant_input)
                    
                    if cantidad <= 0:
                        print("La cantidad debe ser mayor a 0.")
                    elif cantidad <= p['stock']:
                        subtotal = p['precio'] * cantidad
                        total_venta += subtotal
                        
                        carrito.append({
                            'producto_obj': p,
                            'nombre': p['nombre'],
                            'cantidad': cantidad,
                            'subtotal': subtotal
                        })
                        
                        print(f"Agregado: {cantidad}x {p['nombre']} - Subtotal: ${subtotal:.2f}")
                    else:
                        print("Error: Stock insuficiente.")
                except ValueError:
                    print("Error: Ingrese un número válido.")
                break
        
        if not encontrado:
            print("El producto o código no existe.")

    if carrito:
        print(f"\n" + "="*20)
        print(f"TOTAL A COBRAR: ${total_venta:.2f}")
        print("="*20)
        
        print("Métodos de pago: 1. Efectivo | 2. Tarjeta | 3. QR")
        metodo = input("Seleccione una opción: ")
        
        medios = {"1": "Efectivo", "2": "Tarjeta", "3": "Transferencia/QR"}
        pago_elegido = medios.get(metodo, "Otro")

        for item in carrito:
            p = item['producto_obj']

            p['stock'] -= item['cantidad']

            conexion = sqlite3.connect("inventario.db")
            cursor = conexion.cursor()

            cursor.execute("""
                UPDATE productos
                SET stock = ?
                WHERE id = ?
            """, (
                p["stock"],
                p["id"]
            ))


            registrar_movimiento(
                item['nombre'],
                "VENTA",
                -item['cantidad'],
                item['subtotal'],
                pago_elegido
            )
            
            conexion.commit()
            conexion.close()
        
        
        ruta_ticket = generar_ticket(carrito, total_venta, pago_elegido)
        print(f"\nVenta finalizada. PDF generado: {ruta_ticket}")
        
        if input("¿Desea abrir el ticket ahora? (si/no): ").lower() == 'si':
            abrir_archivo(ruta_ticket)
    else:
        print("\nVenta cancelada (carrito vacío).")
        
def abrir_archivo(ruta):
    
    if os.path.exists(ruta):
        try:
            if platform.system() == "Windows":
                os.startfile(ruta)
            elif platform.system() == "Darwin":
                subprocess.Popen(["open", ruta])
            else:
                subprocess.Popen(["xdg-open", ruta])
        except Exception as e:
            print(f"No se pudo abrir el archivo: {e}")
    else:
        print(f"El archivo no existe en: {ruta}")
        
def reponer_stock(productos):
    print("\n>>> Reponer Stock <<<")

    entrada = input("\nProducto a reponer (escanee código o escriba nombre): ").strip()

    if entrada == "0":
        return

    encontrado = False

    for p in productos:
        if p.get('codigo') == entrada or p['nombre'].lower() == entrada.lower():

            encontrado = True

            print(f"\nProducto: {p['nombre']} | Stock actual: {p['stock']}")

            try:
                cantidad = int(input("Cantidad a ingresar: "))

                if cantidad <= 0:
                    print("La cantidad debe ser mayor a 0")

                else:
                    p['stock'] += cantidad

                    conexion = sqlite3.connect("inventario.db")
                    cursor = conexion.cursor()

                    cursor.execute("""
                        UPDATE productos
                        SET stock = ?
                        WHERE id = ?
                    """, (
                        p["stock"],
                        p["id"]
                    ))

                    conexion.commit()
                    conexion.close()

                    registrar_movimiento(
                        p['nombre'],
                        "REPOSICION",
                        f"+{cantidad}"
                    )

                    print(
                        f"Ingreso Exitoso. Nuevo stock: {p['stock']}"
                    )

            except ValueError:
                print("Error: Ingrese un número válido")

            break

    if not encontrado:
        print(
            f"No existe el producto con el código o nombre: {entrada}"
        )
#-----------------------CODIGO VIEJO COMENTADO (Registrar_movimiento)----------------------       
# def registrar_movimiento(producto, tipo, cantidad):
#     nuevo_movimiento = {
#         "fecha": datetime.now().strftime("%d/%m/%Y, %H:%M:%S"),
#         "producto": producto,
#         "tipo": tipo,
#         "cantidad": cantidad
#     }
#     try:
#         with open("historial.json", "r") as archivo:
#             historial = json.load(archivo)
#     except (FileNotFoundError, json.JSONDecodeError):
#         historial = []
        
#     historial.append(nuevo_movimiento)
    
#     with open("historial.json", "w") as archivo:
#         json.dump(historial, archivo, indent=4)
        
def mostrar_historial():
    print("\n" + "="*40)
    print("     HISTORIAL DE MOVIMIENTOS")
    print("=" *40)
    
    try:
        with open("movimientos.json", "r") as archivo:
            historial = json.load(archivo)
            
        if not historial:
            print("\nPor el momento no hay movimientos")
            return
        
        print(f"{'FECHA':<20} | {'TIPO':<12} | {'PRODUCTO':<15} | {'CANT'}")
        print("-" *60)
    
        for m in historial[-10:]:
            print(f"{m['fecha']:<20} | {m['tipo']:<12} | {m['producto']:<15} | {m['cantidad']}")
            
    except FileNotFoundError:
        print("No hay historial registrado aun.")
        
        #CREA EXCEL USANDO PANDAS
def exportar_excel(productos):
    if not productos:
        print("No hay datos para exportar.")
        return

    nombre_archivo = "Inventario_exportado.xlsx"
    df = pd.DataFrame(productos)
    
    try:
        with pd.ExcelWriter(nombre_archivo, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Productos')
            
            workbook = writer.book
            worksheet = writer.sheets['Productos']
            # ------------------------------ ESTILOS DEL EXCEL --------------------------------

            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=12)
            
            center_alignment = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'), 
                top=Side(style='thin'), bottom=Side(style='thin')
            )

            for col_num, value in enumerate(df.columns, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment
                cell.border = thin_border
                
            for i, col in enumerate(df.columns, 1):
                max_len = max(df[df.columns[i-1]].astype(str).map(len).max(), len(df.columns[i-1])) + 4
                worksheet.column_dimensions[chr(64 + i)].width = max_len

                for row_num in range(2, len(df) + 2):
                    cell = worksheet.cell(row=row_num, column=i)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="left")
                    
            # ----------------------------- FIN DEL ESTILO EXCEL ------------------------------------
        
        print(f"Archivo '{nombre_archivo}' generado con éxito.")
        
        abrir = input("¿Desea abrir el archivo ahora? (si/no): ")
        if abrir.lower() == 'si':
            abrir_excel()

    except PermissionError:
        print("\n" + "-"*40)
        print("ERROR DE PERMISOS")
        print(f"\nNo se puede guardar '{nombre_archivo}' porque ya está abierto.")
        print("\nPor favor, cierra el archivo en Excel e intenta de nuevo.")
        print("-"*40)
    except Exception as e:
        print(f"Ocurrió un error inesperado: {e}")
    
def abrir_excel():
    nombre_archivo = "Inventario_exportado.xlsx"
    
    if os.path.exists(nombre_archivo):
        try:
            if platform.system() == "Windows":
                os.startfile(nombre_archivo)
            elif platform.system() == "Darwin":
                subprocess.Popen(["open", nombre_archivo])
            else:
                subprocess.Popen(["xdg-open", nombre_archivo])
        except Exception as e:
            print(f"No se pudo abrir el archivo: {e}")
    else:
        print("Aún no se ha generado ningún reporte Excel.")
        
    #ESTO CREE PARA LOS BACKUPS SI SE BORRAN LOS ARCHIVOS INCORRECTOS   
def crear_backup(manual=False):
    if not os.path.exists("backups"):
        os.makedirs("backups")

    archivo_original = "inventario.db"

    if os.path.exists(archivo_original):

        fecha_str = datetime.now().strftime("%d%m%Y_%H%M")
        nombre_backup = f"backups/backup_{fecha_str}.db"

        try:
            shutil.copy2(archivo_original, nombre_backup)

            if manual:
                print(f"\nRespaldo creado: {nombre_backup}")

        except Exception as e:

            if manual:
                print(f"\nError al crear backup: {e}")

    try:
        lista_backups = sorted(
            [os.path.join("backups", f) for f in os.listdir("backups")],
            key=os.path.getmtime
        )

        while len(lista_backups) > 10:
            os.remove(lista_backups.pop(0))

    except Exception:
        pass
    
def abrir_backups():
    ruta = os.path.abspath("backups")
    
    if not os.path.exists(ruta):
        print("La carpeta backup no ha sido creada")
        return
    try:
        if platform.system() == "Windows":
            os.startfile(ruta)
        elif platform.system() == "Darwin":
            subprocess.Popen(["open", ruta])
        else:
            subprocess.Popen(["xdg-open", ruta])
    except Exception as e:
        print(f"No se puede abrir: {e}")
        
    #CREE UN GENERADOR DE TICKET FANTASIA
def generar_ticket(carrito, total, pago_elegido):
    if not os.path.exists("tickets"):
        os.makedirs("tickets", exist_ok=True)
    
    num_ticket = len(os.listdir("tickets")) + 1
    pdf = FPDF(format=(80, 150)) 
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=5)
    
    pdf.set_font("Arial", "B", 12)
    pdf.cell(0, 10, "MI NEGOCIO S.A.", ln=True, align="C")
    pdf.set_font("Arial", "", 8)
    pdf.cell(0, 5, "AV. SIEMPRE VIVA 123", ln=True, align="C")
    pdf.cell(0, 5, f"Ticket Nro: {num_ticket:04d}", ln=True, align="C")
    pdf.cell(0, 5, f"Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M')}", ln=True, align="C")
    pdf.ln(5)
    
    pdf.set_font("Arial", "B", 8)
    pdf.cell(30, 5, "Producto", border="B")
    pdf.cell(10, 5, "Cant", border="B", align="C")
    pdf.cell(20, 5, "Total", border="B", align="R")
    pdf.ln(6)
    
    pdf.set_font("Arial", "", 8)
    for p in carrito:
        pdf.cell(30, 5, p['nombre'][:15])
        pdf.cell(10, 5, str(p['cantidad']), align="C")
        pdf.cell(20, 5, f"${p['subtotal']:.2f}", align="R")
        pdf.ln(5)
    
    pdf.ln(5)
    pdf.set_font("Arial", "B", 10)
    pdf.cell(40, 10, "TOTAL FINAL:", border="T")
    pdf.cell(20, 10, f"${total:.2f}", border="T", align="R")
    pdf.ln(10)
    
    pdf.set_font("Arial", "I", 8)
    pdf.cell(0, 5, f"Metodo de Pago: {pago_elegido}", ln=True, align="L")
    pdf.cell(0, 10, "¡Gracias por su compra!", ln=True, align="C")
    
    nombre_pdf = os.path.join("tickets", f"ticket_{num_ticket:04d}.pdf")
    pdf.output(nombre_pdf)
    return nombre_pdf

        #HECHO PARA ELIMINAR LOS TICKETS MAYORES A UNA SEMANA
def limpiar_tickets_viejos(dias=7):
    ruta_tickets = "tickets"
    if not os.path.exists(ruta_tickets):
        return

    ahora = time.time()
    segundos_limite = dias * 24 * 60 * 60
    
    archivos_borrados = 0
    
    for archivo in os.listdir(ruta_tickets):
        ruta_completa = os.path.join(ruta_tickets, archivo)
        
        if os.path.isfile(ruta_completa):
            fecha_archivo = os.path.getmtime(ruta_completa)
            
            if (ahora - fecha_archivo) > segundos_limite:
                try:
                    os.remove(ruta_completa)
                    archivos_borrados += 1
                except Exception as e:
                    print(f"No se pudo borrar {archivo}: {e}")
    
    if archivos_borrados > 0:
        print(f"Limpieza: Se eliminaron {archivos_borrados} tickets antiguos.")
        
def registrar_movimiento(producto, tipo, cantidad, monto=0, metodo="Efectivo"):
    nuevo = {
        "fecha": datetime.now().strftime("%d-%m-%Y %H:%M"),
        "producto": producto,
        "tipo": tipo,
        "cantidad": cantidad,
        "monto": monto,
        "metodo": metodo
        
    }
    try:
        try:
            with open("movimientos.json", "r") as f:
                historial = json.load(f)
        except (FileNotFoundError, json.JSONDecodeError):
            historial = []
        
        historial.append(nuevo)
    
        with open("movimientos.json", "w") as f:
            json.dump(historial, f, indent=4)
    except Exception as e:
        print(f"Error al registrar historial: {e}")
        
def caja_diaria():
    print("\n" + "="*35)
    print("      RESUMEN DE CAJA DIARIO")
    print("="*35)
    
    hoy = datetime.now().strftime("%d-%m-%Y")
    totales = {"Efectivo": 0, "Tarjeta": 0, "Transferencia/QR": 0}
    ventas_cont = 0

    try:
        with open("movimientos.json", "r") as f:
            movimientos = json.load(f)
            
        for m in movimientos:
            if m['fecha'].startswith(hoy) and m['tipo'] == "VENTA":
                metodo = m.get('metodo', 'Efectivo')
                monto = m.get('monto', 0)
                
                if metodo in totales:
                    totales[metodo] += monto
                ventas_cont += 1

        if ventas_cont > 0:
            print(f" Fecha: {hoy} |  Ventas: {ventas_cont}")
            print("-" * 35)
            for medio, suma in totales.items():
                print(f" {medio:<18}: ${suma:>10.2f}")
            print("-" * 35)
            print(f"TOTAL RECAUDADO:  ${sum(totales.values()):>10.2f}")
        else:
            print(f"\nNo se encontraron ventas hoy ({hoy}).")
            
    except FileNotFoundError:
        print("\nAún no hay movimientos registrados.")
    
    print("\n" + "-" * 20)
    input("Enter para volver al menú")
    return